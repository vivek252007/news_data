#!/usr/bin/python
# -*- coding: utf-8 -*-
from openpyxl import Workbook,load_workbook
import os.path,sys

class excel_data():
	def __init__(self,ticker):
		self.wb = Workbook()
		self.worksheet = 0
		self.ticker = ticker.upper()
		self.dir_name = '/home/vivek/Desktop/Data/News_Data/'
		reload(sys)
		sys.setdefaultencoding('UTF8')

	def open_sheet(self):
		if not (os.path.isfile(self.dir_name+self.ticker+'.xlsx')):
			self.wb = Workbook()
			self.worksheet = self.wb.create_sheet(0)
			self.worksheet.title = self.ticker+" NEWS"
			tags = ['Date','Time','Stock Price','Headline','Description','Link']
			self.worksheet.append(tags)
		else:
			self.wb = load_workbook(self.dir_name+self.ticker+'.xlsx')
			self.worksheet = self.wb[self.ticker+" NEWS"]
 
	def write_sheet(self,data):
		chk_count =  (self.worksheet.max_row)
		print chk_count
		if (chk_count -60) <=0:
			chk_start = 1
		else :
			chk_start = chk_count-60
		wrt_token = True
		for row in self.worksheet.iter_rows('D'+str(chk_start)+':D'+str(chk_count)+''):
			for cell in row:
				if data[3] == cell.value:
					wrt_token = False 
					break

		if (wrt_token):
			print "**** "+data[3]+" ****"
			self.worksheet.append(data)
			return data[3]
		else :
			return False

	def reading_sheet(self):
		chk_count =  (self.worksheet.max_row)
		output_data = []
		for i in range(2,chk_count+1):
			output_data.append(self.worksheet['D'+str(i)+''].value)
		return output_data

	def close_sheet(self):
		self.wb.save(self.dir_name+self.ticker+'.xlsx')


if __name__=="__main__":
	obj = excel_data('GOOG')
	data = ['07-06-2016','9:12:36','25.36','This is amazing','All of it is amazing','htttp://visit.me']
	obj.write_sheet(data)